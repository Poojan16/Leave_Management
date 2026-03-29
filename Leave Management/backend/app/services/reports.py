import io
from datetime import date
from typing import Any

# ── PDF ───────────────────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ── Excel ─────────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────

_COLUMNS = [
    ("Employee",    "employee",    5 * cm),
    ("Dept",        "department",  3 * cm),
    ("Type",        "leave_type",  3 * cm),
    ("Start",       "start_date",  2.8 * cm),
    ("End",         "end_date",    2.8 * cm),
    ("Days",        "days",        1.5 * cm),
    ("Status",      "status",      2.5 * cm),
    ("Approved By", "approved_by", 4 * cm),
]

_STATUS_COLORS = {
    "approved":  colors.HexColor("#d1fae5"),
    "rejected":  colors.HexColor("#fee2e2"),
    "pending":   colors.HexColor("#fef9c3"),
    "cancelled": colors.HexColor("#f3f4f6"),
}

_HEADER_BG   = colors.HexColor("#1e3a5f")
_HEADER_FG   = colors.white
_ALT_ROW_BG  = colors.HexColor("#f0f4f8")
_BRAND_COLOR = colors.HexColor("#1e3a5f")


def _filter_summary(filters: dict[str, Any]) -> str:
    parts = []
    if filters.get("year"):
        parts.append(f"Year: {filters['year']}")
    if filters.get("month"):
        import calendar
        parts.append(f"Month: {calendar.month_name[filters['month']]}")
    if filters.get("status"):
        parts.append(f"Status: {filters['status'].capitalize()}")
    if filters.get("dept_id"):
        parts.append(f"Dept ID: {filters['dept_id']}")
    return "  |  ".join(parts) if parts else "All records"


# ── PDF ───────────────────────────────────────────────────────────────────────

def generate_pdf_report(data: list[dict[str, Any]], filters: dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=_BRAND_COLOR,
        spaceAfter=4,
        alignment=TA_LEFT,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#6b7280"),
        spaceAfter=12,
    )
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8)

    # ── Header rows ───────────────────────────────────────────────────────────
    col_headers = [Paragraph(f"<b>{h}</b>", ParagraphStyle(
        "H", parent=styles["Normal"], fontSize=8, textColor=_HEADER_FG
    )) for h, _, _ in _COLUMNS]

    table_data = [col_headers]
    row_styles: list[tuple] = []

    for i, row in enumerate(data, start=1):
        status = str(row.get("status", "")).lower()
        bg = _STATUS_COLORS.get(status, colors.white)
        if i % 2 == 0:
            bg = _ALT_ROW_BG if status not in _STATUS_COLORS else bg

        cells = [
            Paragraph(str(row.get(key, "")), cell_style)
            for _, key, _ in _COLUMNS
        ]
        table_data.append(cells)
        row_styles.append(("BACKGROUND", (0, i), (-1, i), bg))

    col_widths = [w for _, _, w in _COLUMNS]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    base_style = [
        ("BACKGROUND",  (0, 0), (-1, 0), _HEADER_BG),
        ("TEXTCOLOR",   (0, 0), (-1, 0), _HEADER_FG),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 8),
        ("ROWBACKGROUND", (0, 1), (-1, -1), [colors.white, _ALT_ROW_BG]),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]
    table.setStyle(TableStyle(base_style + row_styles))

    def _add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#9ca3af"))
        canvas.drawString(1.5 * cm, 1 * cm, f"Generated: {date.today()}")
        canvas.drawRightString(
            landscape(A4)[0] - 1.5 * cm, 1 * cm,
            f"Page {doc.page}"
        )
        canvas.restoreState()

    story = [
        Paragraph("Leave Management System — Leave Report", title_style),
        Paragraph(_filter_summary(filters), subtitle_style),
        Paragraph(f"Total records: {len(data)}", subtitle_style),
        Spacer(1, 0.3 * cm),
        table,
    ]

    doc.build(story, onFirstPage=_add_page_number, onLaterPages=_add_page_number)
    return buffer.getvalue()


# ── Excel ─────────────────────────────────────────────────────────────────────

def generate_excel_report(data: list[dict[str, Any]], filters: dict[str, Any]) -> bytes:
    wb = Workbook()

    # ── Data sheet ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Leave Report"

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1e3a5f")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    alt_fill  = PatternFill("solid", fgColor="F0F4F8")
    status_fills = {
        "approved":  PatternFill("solid", fgColor="D1FAE5"),
        "rejected":  PatternFill("solid", fgColor="FEE2E2"),
        "pending":   PatternFill("solid", fgColor="FEF9C3"),
        "cancelled": PatternFill("solid", fgColor="F3F4F6"),
    }

    headers = [h for h, _, _ in _COLUMNS]
    ws.append(headers)

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font   = header_font
        cell.fill   = header_fill
        cell.border = border
        cell.alignment = header_align

    ws.row_dimensions[1].height = 22

    for row_idx, row in enumerate(data, start=2):
        values = [str(row.get(key, "")) for _, key, _ in _COLUMNS]
        ws.append(values)

        status = str(row.get("status", "")).lower()
        fill = status_fills.get(status, PatternFill("solid", fgColor="FFFFFF"))
        if row_idx % 2 == 0 and status not in status_fills:
            fill = alt_fill

        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Auto column widths
    for col_idx, (header, key, _) in enumerate(_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in data:
            val = str(row.get(key, ""))
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15

    summary_header_font = Font(bold=True, color="FFFFFF", size=12)
    summary_header_fill = PatternFill("solid", fgColor="1e3a5f")

    ws2.append(["Leave Report Summary"])
    ws2["A1"].font  = Font(bold=True, size=14, color="1e3a5f")
    ws2["A1"].alignment = Alignment(horizontal="left")
    ws2.append([])
    ws2.append(["Filter", "Value"])
    for cell in ws2[3]:
        cell.font   = summary_header_font
        cell.fill   = summary_header_fill
        cell.border = border

    for k, v in filters.items():
        if v is not None:
            ws2.append([k.replace("_", " ").title(), str(v)])

    ws2.append([])
    ws2.append(["Metric", "Count"])
    for cell in ws2[ws2.max_row]:
        cell.font   = summary_header_font
        cell.fill   = summary_header_fill
        cell.border = border

    from collections import Counter
    status_counts = Counter(str(r.get("status", "")).lower() for r in data)
    ws2.append(["Total Records", len(data)])
    for s, cnt in sorted(status_counts.items()):
        ws2.append([s.capitalize(), cnt])

    total_days = sum(int(r.get("days", 0)) for r in data if str(r.get("status", "")).lower() == "approved")
    ws2.append(["Total Approved Days", total_days])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Stats dict ────────────────────────────────────────────────────────────────

def generate_stats_report(stats: dict[str, Any]) -> dict[str, Any]:
    """Return stats dict formatted for frontend chart consumption."""
    return {
        "summary": {
            "year": stats.get("year"),
            "total_employees": stats.get("total_employees", 0),
            "total_requests": stats.get("total_requests", 0),
            "pending": stats.get("pending", 0),
            "approved": stats.get("approved", 0),
            "rejected": stats.get("rejected", 0),
            "cancelled": stats.get("cancelled", 0),
            "total_days_taken": stats.get("total_days_taken", 0),
            "approval_rate": round(
                stats.get("approved", 0) / stats.get("total_requests", 1) * 100, 1
            ) if stats.get("total_requests") else 0,
        },
        "charts": {
            "by_status": [
                {"label": "Pending",   "value": stats.get("pending",   0), "color": "#f59e0b"},
                {"label": "Approved",  "value": stats.get("approved",  0), "color": "#10b981"},
                {"label": "Rejected",  "value": stats.get("rejected",  0), "color": "#ef4444"},
                {"label": "Cancelled", "value": stats.get("cancelled", 0), "color": "#6b7280"},
            ],
            "by_department": stats.get("by_department", []),
            "by_leave_type": stats.get("by_leave_type", []),
            "by_month": stats.get("by_month", []),
        },
    }
